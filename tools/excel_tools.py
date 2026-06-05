import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from sandbox import resolve_path
import os

def escribir_excel(filename: str, hoja: str, datos: list, context: str = "outputs") -> str:
    """
    Crea un nuevo archivo Excel o sobrescribe si existe.
    - datos debe ser una lista de listas: [[header1, header2], [val1, val2]]
    """
    try:
        filepath = resolve_path(filename, context)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = hoja

        for fila in datos:
            ws.append(fila)

        wb.save(filepath)
        return f"Excel escrito con éxito: {filepath}"
    except Exception as e:
        return f"Error escribiendo Excel: {e}"

def actualizar_excel_existente(filename: str, hoja: str, datos: list, context: str = "outputs") -> str:
    """Abre un Excel existente y añade filas al final."""
    try:
        filepath = resolve_path(filename, context)
        if not os.path.exists(filepath):
            return f"Error: No se encontró el archivo {filepath}"

        wb = openpyxl.load_workbook(filepath)
        if hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            ws = wb.create_sheet(hoja)

        for fila in datos:
            ws.append(fila)

        wb.save(filepath)
        return f"Excel actualizado con éxito: {filepath}"
    except Exception as e:
        return f"Error actualizando Excel: {e}"

def aplicar_formato_excel(filename: str, context: str = "outputs") -> str:
    """
    Aplica formato básico a un Excel:
    - Encabezados en negrita con fondo
    - Alineación centrada
    - Autofit aproximado de columnas
    """
    try:
        filepath = resolve_path(filename, context)
        if not os.path.exists(filepath):
            return f"Error: No se encontró el archivo {filepath}"

        wb = openpyxl.load_workbook(filepath)
        for hoja_nombre in wb.sheetnames:
            ws = wb[hoja_nombre]

            # Formatear encabezado (primera fila)
            fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Autofit simple (ajuste de ancho de columnas por contenido)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    if cell.value is None:
                        continue
                    longitud = len(str(cell.value))  # str() para soportar números
                    if longitud > max_length:
                        max_length = longitud
                # Limitar el ancho máximo para que columnas largas no desborden
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column].width = adjusted_width

        nuevo_nombre = os.path.splitext(filename)[0] + "_formateado.xlsx"
        nuevo_filepath = resolve_path(nuevo_nombre, context)
        wb.save(nuevo_filepath)
        return f"Excel formateado con éxito. Guardado como: {nuevo_filepath}"
    except Exception as e:
        return f"Error formateando Excel: {e}"

import json
def crear_excel_complejo_desde_json(filename: str, json_str: str, context: str = "outputs") -> str:
    """
    Crea un archivo Excel complejo a partir de un JSON estructurado de ChatGPT.
    Estructura esperada (Lista de hojas, cada hoja tiene un nombre y filas):
    [
        {
            "hoja": "Resumen Financiero",
            "datos": [
                ["Mes", "Ingresos", "Egresos", "Balance"],
                ["Enero", 15000, 10000, 5000],
                ["Febrero", 16000, 12000, 4000]
            ]
        },
        ...
    ]
    """
    try:
        data = json.loads(json_str)
        if not data:
            return "Error: JSON vacío o mal formado."

        filepath = resolve_path(filename, context)
        wb = openpyxl.Workbook()
        default_sheet = wb.active

        hojas_creadas = 0
        for idx, hoja_datos in enumerate(data):
            if not isinstance(hoja_datos, dict):
                continue
            # Nombre de hoja saneado (Excel limita a 31 chars y prohíbe : \ / ? * [ ])
            nombre_hoja = str(hoja_datos.get("hoja", f"Hoja{idx+1}"))
            for c in r':\/?*[]':
                nombre_hoja = nombre_hoja.replace(c, " ")
            nombre_hoja = nombre_hoja.strip()[:31] or f"Hoja{idx+1}"

            filas = hoja_datos.get("datos", [])
            ws = wb.create_sheet(nombre_hoja)
            for fila in filas:
                if isinstance(fila, list):
                    ws.append(fila)
            hojas_creadas += 1

        # Guarda: si no se creó ninguna hoja válida, no romper openpyxl
        if hojas_creadas == 0:
            default_sheet.append(["Sin datos generados"])
        else:
            wb.remove(default_sheet)

        wb.save(filepath)
        # Aplicamos formato automáticamente a todas las hojas del archivo recién creado
        aplicar_formato_excel(filename, context)

        return f"Archivo Excel Complejo creado y formateado con éxito: {filepath}"

    except json.JSONDecodeError as e:
         return f"Error parseando JSON de Excel: {e}\nJSON Recibido:\n{json_str[:500]}"
    except Exception as e:
        return f"Error creando Excel complejo: {e}"
